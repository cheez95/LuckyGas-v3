"""
Data export service for Lucky Gas.
Handles exporting data to various formats (CSV, Excel, JSON).
"""
import io
import csv
import json
import zipfile
from datetime import datetime, date
from typing import Dict, Any, List, Optional, BinaryIO
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_
import aiofiles

from app.models.customer import Customer
from app.models.order import Order
from app.models.order_item import OrderItem
from app.models.route import Route, RouteStop
from app.models.driver import Driver
from app.models.gas_product import GasProduct as Product
from app.models.delivery import Delivery
from app.models.invoice import Invoice, InvoiceItem
# Payment model imported from order module if needed
from app.core.config import settings
from app.utils.datetime_utils import format_taiwan_date

import logging

logger = logging.getLogger(__name__)


class DataExportService:
    """Service for exporting data in various formats."""
    
    def __init__(self, db: AsyncSession):
        self.db = db
        
    async def export_customers(
        self,
        format: str = "csv",
        filters: Optional[Dict[str, Any]] = None,
        include_fields: Optional[List[str]] = None
    ) -> BinaryIO:
        """Export customer data."""
        # Build query
        query = select(Customer)
        
        if filters:
            if "customer_type" in filters:
                query = query.where(Customer.customer_type == filters["customer_type"])
            if "district" in filters:
                query = query.where(Customer.district == filters["district"])
            if "is_active" in filters:
                query = query.where(Customer.is_active == filters["is_active"])
        
        result = await self.db.execute(query)
        customers = result.scalars().all()
        
        # Convert to DataFrame
        data = []
        for customer in customers:
            row = {
                "客戶編號": customer.id,
                "客戶名稱": customer.name,
                "電話": customer.phone,
                "地址": customer.address,
                "區域": customer.district,
                "客戶類型": customer.customer_type,
                "信用額度": customer.credit_limit,
                "付款條件": customer.payment_terms,
                "聯絡人": customer.contact_person,
                "聯絡人電話": customer.contact_phone,
                "電子郵件": customer.email,
                "統一編號": customer.tax_id,
                "狀態": "啟用" if customer.is_active else "停用",
                "建立日期": format_taiwan_date(customer.created_at),
                "備註": customer.notes
            }
            
            # Filter fields if specified
            if include_fields:
                row = {k: v for k, v in row.items() if k in include_fields}
            
            data.append(row)
        
        df = pd.DataFrame(data)
        
        # Export based on format
        if format == "csv":
            return self._export_to_csv(df, "customers")
        elif format == "excel":
            return self._export_to_excel({"客戶資料": df}, "customers")
        elif format == "json":
            return self._export_to_json(data, "customers")
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    async def export_orders(
        self,
        format: str = "csv",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        status: Optional[str] = None,
        customer_id: Optional[int] = None
    ) -> BinaryIO:
        """Export order data with items."""
        # Build query
        query = select(Order).options(
            selectinload(Order.customer),
            selectinload(Order.items).selectinload(OrderItem.product)
        )
        
        conditions = []
        if start_date:
            conditions.append(Order.delivery_date >= start_date)
        if end_date:
            conditions.append(Order.delivery_date <= end_date)
        if status:
            conditions.append(Order.status == status)
        if customer_id:
            conditions.append(Order.customer_id == customer_id)
        
        if conditions:
            query = query.where(and_(*conditions))
        
        result = await self.db.execute(query)
        orders = result.scalars().unique().all()
        
        # Convert to DataFrame
        order_data = []
        item_data = []
        
        for order in orders:
            order_row = {
                "訂單編號": order.id,
                "客戶名稱": order.customer.name,
                "配送日期": format_taiwan_date(order.delivery_date),
                "配送地址": order.delivery_address,
                "狀態": self._translate_order_status(order.status),
                "總金額": order.total_amount,
                "折扣": order.discount_amount,
                "稅額": order.tax_amount,
                "應付金額": order.final_amount,
                "付款方式": order.payment_method,
                "司機": order.driver.name if order.driver else "",
                "路線": order.route_id,
                "建立時間": format_taiwan_date(order.created_at),
                "備註": order.notes
            }
            order_data.append(order_row)
            
            # Order items
            for item in order.items:
                item_row = {
                    "訂單編號": order.id,
                    "產品名稱": item.product.name,
                    "數量": item.quantity,
                    "單價": item.unit_price,
                    "小計": item.subtotal,
                    "折扣": item.discount_amount,
                    "總計": item.total_amount
                }
                item_data.append(item_row)
        
        orders_df = pd.DataFrame(order_data)
        items_df = pd.DataFrame(item_data)
        
        # Export based on format
        if format == "csv":
            # Create zip with both files
            return self._export_multiple_csv({
                "orders": orders_df,
                "order_items": items_df
            }, "orders")
        elif format == "excel":
            return self._export_to_excel({
                "訂單": orders_df,
                "訂單明細": items_df
            }, "orders")
        elif format == "json":
            # Combine data
            combined_data = []
            for order in orders:
                order_dict = {
                    "id": order.id,
                    "customer": order.customer.name,
                    "delivery_date": order.delivery_date.isoformat(),
                    "status": order.status,
                    "total_amount": float(order.total_amount),
                    "items": [
                        {
                            "product": item.product.name,
                            "quantity": item.quantity,
                            "unit_price": float(item.unit_price),
                            "total": float(item.total_amount)
                        }
                        for item in order.items
                    ]
                }
                combined_data.append(order_dict)
            return self._export_to_json(combined_data, "orders")
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    async def export_routes(
        self,
        format: str = "csv",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        driver_id: Optional[int] = None
    ) -> BinaryIO:
        """Export route data with stops."""
        # Build query
        query = select(Route).options(
            selectinload(Route.driver),
            selectinload(Route.stops).selectinload(RouteStop.order).selectinload(Order.customer)
        )
        
        conditions = []
        if start_date:
            conditions.append(Route.date >= start_date)
        if end_date:
            conditions.append(Route.date <= end_date)
        if driver_id:
            conditions.append(Route.driver_id == driver_id)
        
        if conditions:
            query = query.where(and_(*conditions))
        
        result = await self.db.execute(query)
        routes = result.scalars().unique().all()
        
        # Convert to DataFrame
        route_data = []
        stop_data = []
        
        for route in routes:
            route_row = {
                "路線編號": route.id,
                "日期": format_taiwan_date(route.date),
                "司機": route.driver.name if route.driver else "",
                "狀態": self._translate_route_status(route.status),
                "總站點": len(route.stops),
                "總距離(km)": route.total_distance_km,
                "預計時間(分)": route.estimated_duration_minutes,
                "開始時間": format_taiwan_date(route.start_time) if route.start_time else "",
                "結束時間": format_taiwan_date(route.end_time) if route.end_time else "",
                "建立時間": format_taiwan_date(route.created_at)
            }
            route_data.append(route_row)
            
            # Route stops
            for stop in route.stops:
                stop_row = {
                    "路線編號": route.id,
                    "順序": stop.sequence_number,
                    "客戶": stop.order.customer.name if stop.order else "",
                    "地址": stop.address,
                    "預計到達": format_taiwan_date(stop.estimated_arrival) if stop.estimated_arrival else "",
                    "實際到達": format_taiwan_date(stop.actual_arrival) if stop.actual_arrival else "",
                    "狀態": self._translate_stop_status(stop.status),
                    "備註": stop.notes
                }
                stop_data.append(stop_row)
        
        routes_df = pd.DataFrame(route_data)
        stops_df = pd.DataFrame(stop_data)
        
        # Export based on format
        if format == "csv":
            return self._export_multiple_csv({
                "routes": routes_df,
                "route_stops": stops_df
            }, "routes")
        elif format == "excel":
            return self._export_to_excel({
                "路線": routes_df,
                "路線站點": stops_df
            }, "routes")
        elif format == "json":
            combined_data = []
            for route in routes:
                route_dict = {
                    "id": route.id,
                    "date": route.date.isoformat(),
                    "driver": route.driver.name if route.driver else None,
                    "status": route.status,
                    "stops": [
                        {
                            "sequence": stop.sequence_number,
                            "customer": stop.order.customer.name if stop.order else None,
                            "address": stop.address,
                            "status": stop.status
                        }
                        for stop in sorted(route.stops, key=lambda x: x.sequence_number)
                    ]
                }
                combined_data.append(route_dict)
            return self._export_to_json(combined_data, "routes")
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    async def export_deliveries(
        self,
        format: str = "csv",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> BinaryIO:
        """Export delivery history."""
        query = select(Delivery).options(
            selectinload(Delivery.order).selectinload(Order.customer),
            selectinload(Delivery.driver)
        )
        
        conditions = []
        if start_date:
            conditions.append(Delivery.delivered_at >= start_date)
        if end_date:
            conditions.append(Delivery.delivered_at <= end_date)
        
        if conditions:
            query = query.where(and_(*conditions))
        
        result = await self.db.execute(query)
        deliveries = result.scalars().all()
        
        # Convert to DataFrame
        data = []
        for delivery in deliveries:
            row = {
                "配送編號": delivery.id,
                "訂單編號": delivery.order_id,
                "客戶": delivery.order.customer.name,
                "司機": delivery.driver.name,
                "配送時間": format_taiwan_date(delivery.delivered_at),
                "簽名": "有" if delivery.signature_image else "無",
                "照片": "有" if delivery.photo_image else "無",
                "瓦斯桶序號": delivery.cylinder_serial_number,
                "備註": delivery.notes
            }
            data.append(row)
        
        df = pd.DataFrame(data)
        
        if format == "csv":
            return self._export_to_csv(df, "deliveries")
        elif format == "excel":
            return self._export_to_excel({"配送記錄": df}, "deliveries")
        elif format == "json":
            return self._export_to_json(data, "deliveries")
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    async def export_financial_summary(
        self,
        format: str = "excel",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> BinaryIO:
        """Export financial summary report."""
        # Get invoices
        invoice_query = select(Invoice).options(
            selectinload(Invoice.customer),
            selectinload(Invoice.items).selectinload(InvoiceItem.product)
        )
        
        if start_date and end_date:
            invoice_query = invoice_query.where(
                and_(Invoice.issue_date >= start_date, Invoice.issue_date <= end_date)
            )
        
        invoice_result = await self.db.execute(invoice_query)
        invoices = invoice_result.scalars().unique().all()
        
        # Get payments
        payment_query = select(Payment).options(selectinload(Payment.customer))
        
        if start_date and end_date:
            payment_query = payment_query.where(
                and_(Payment.payment_date >= start_date, Payment.payment_date <= end_date)
            )
        
        payment_result = await self.db.execute(payment_query)
        payments = payment_result.scalars().all()
        
        # Create summary data
        invoice_data = []
        for invoice in invoices:
            invoice_data.append({
                "發票號碼": invoice.invoice_number,
                "客戶": invoice.customer.name,
                "開立日期": format_taiwan_date(invoice.issue_date),
                "到期日": format_taiwan_date(invoice.due_date),
                "總金額": invoice.total_amount,
                "已付金額": invoice.paid_amount,
                "未付金額": invoice.total_amount - invoice.paid_amount,
                "狀態": self._translate_invoice_status(invoice.status)
            })
        
        payment_data = []
        for payment in payments:
            payment_data.append({
                "付款編號": payment.id,
                "客戶": payment.customer.name,
                "付款日期": format_taiwan_date(payment.payment_date),
                "金額": payment.amount,
                "付款方式": payment.payment_method,
                "參考號碼": payment.reference_number,
                "備註": payment.notes
            })
        
        # Calculate summary
        total_invoiced = sum(i.total_amount for i in invoices)
        total_paid = sum(p.amount for p in payments)
        total_outstanding = sum(i.total_amount - i.paid_amount for i in invoices)
        
        summary_data = [{
            "項目": "應收帳款總額",
            "金額": total_invoiced
        }, {
            "項目": "已收款項",
            "金額": total_paid
        }, {
            "項目": "未收款項",
            "金額": total_outstanding
        }]
        
        invoices_df = pd.DataFrame(invoice_data)
        payments_df = pd.DataFrame(payment_data)
        summary_df = pd.DataFrame(summary_data)
        
        if format == "excel":
            return self._export_to_excel({
                "財務摘要": summary_df,
                "發票明細": invoices_df,
                "付款記錄": payments_df
            }, "financial_summary")
        else:
            raise ValueError(f"Financial summary only supports Excel format")
    
    def _export_to_csv(self, df: pd.DataFrame, filename: str) -> BinaryIO:
        """Export DataFrame to CSV."""
        output = io.BytesIO()
        
        # Write BOM for Excel compatibility with Chinese characters
        output.write(b'\xef\xbb\xbf')
        
        # Convert to CSV
        csv_data = df.to_csv(index=False, encoding='utf-8')
        output.write(csv_data.encode('utf-8'))
        
        output.seek(0)
        return output
    
    def _export_multiple_csv(self, dataframes: Dict[str, pd.DataFrame], filename: str) -> BinaryIO:
        """Export multiple DataFrames to a ZIP file containing CSVs."""
        output = io.BytesIO()
        
        with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zf:
            for name, df in dataframes.items():
                csv_buffer = io.StringIO()
                df.to_csv(csv_buffer, index=False)
                csv_data = csv_buffer.getvalue()
                
                # Add BOM for Excel compatibility
                csv_bytes = b'\xef\xbb\xbf' + csv_data.encode('utf-8')
                zf.writestr(f"{name}.csv", csv_bytes)
        
        output.seek(0)
        return output
    
    def _export_to_excel(self, sheets: Dict[str, pd.DataFrame], filename: str) -> BinaryIO:
        """Export multiple DataFrames to Excel with formatting."""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get the worksheet
                worksheet = writer.sheets[sheet_name]
                
                # Format headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Auto-adjust column widths
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value or "")) for cell in column_cells)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
                
                # Add borders and alignment
                from openpyxl.styles import Border, Side
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(min_row=1):
                    for cell in row:
                        cell.border = thin_border
                        if cell.row > 1:  # Not header
                            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        output.seek(0)
        return output
    
    def _export_to_json(self, data: List[Dict[str, Any]], filename: str) -> BinaryIO:
        """Export data to JSON."""
        output = io.BytesIO()
        
        json_data = json.dumps(data, ensure_ascii=False, indent=2, default=str)
        output.write(json_data.encode('utf-8'))
        
        output.seek(0)
        return output
    
    def _translate_order_status(self, status: str) -> str:
        """Translate order status to Chinese."""
        translations = {
            "pending": "待處理",
            "confirmed": "已確認",
            "preparing": "準備中",
            "delivering": "配送中",
            "delivered": "已送達",
            "cancelled": "已取消",
            "failed": "配送失敗"
        }
        return translations.get(status, status)
    
    def _translate_route_status(self, status: str) -> str:
        """Translate route status to Chinese."""
        translations = {
            "draft": "草稿",
            "planned": "已規劃",
            "published": "已發布",
            "in_progress": "進行中",
            "completed": "已完成",
            "cancelled": "已取消"
        }
        return translations.get(status, status)
    
    def _translate_stop_status(self, status: str) -> str:
        """Translate stop status to Chinese."""
        translations = {
            "pending": "待配送",
            "completed": "已完成",
            "failed": "失敗",
            "skipped": "跳過"
        }
        return translations.get(status, status)
    
    def _translate_invoice_status(self, status: str) -> str:
        """Translate invoice status to Chinese."""
        translations = {
            "draft": "草稿",
            "issued": "已開立",
            "paid": "已付款",
            "partial": "部分付款",
            "overdue": "逾期",
            "cancelled": "已作廢"
        }
        return translations.get(status, status)